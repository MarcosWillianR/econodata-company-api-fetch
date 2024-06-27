import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

class WebScraperUtils:
    def __init__(self):
        self.file = 'empresas_econodata.xlsx'

    def adjust_col_width(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    def cnpj_exists(self, cnpj):
        if os.path.exists(self.file):
            df = pd.read_excel(self.file)
            df['cnpj'] = df['cnpj'].astype(str).str.strip()
            cnpj = str(cnpj).strip()
            return cnpj in df['cnpj'].values
        return False

    def xlsx_size(self):
        if os.path.exists(self.file):
            df = pd.read_excel(self.file)
            return len(df)
        return 0

    def xlsx_update(self, data):
        if os.path.exists(self.file):
            df = pd.read_excel(self.file)
        else:
            df = pd.DataFrame(columns=['empresa', 'cnpj', 'razao_social', 'capital_social', 'telefones', 'pessoas'])

        df_new = pd.DataFrame(data)

        df = df.dropna(how='all', axis=1)
        df_new = df_new.dropna(how='all', axis=1)

        df_updated = pd.concat([df, df_new], ignore_index=True)
        df_updated = df_updated.sort_values(by='capital_social', ascending=False, na_position='last')
        df_updated.to_excel(self.file, index=False)

        wb = load_workbook(self.file)
        ws = wb.active

        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.border = border

        self.adjust_col_width(ws)

        wb.save(self.file)